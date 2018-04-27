# coding=utf8
import os
import datetime
import time
from selenium import webdriver
#from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.common.keys import Keys
import selenium.webdriver.support.ui as ui  
from selenium.webdriver.support.ui import Select ##HTML Selection을 처리하기 위함
from openpyxl import load_workbook #Excel파일을 처리하기위 한 모듈
import telegramSend
import json
from collections import OrderedDict

if __name__ == "__main__": 
   with open('./configData.json', encoding="utf-8") as data_file:    
      data = json.load(data_file, object_pairs_hook=OrderedDict)

   def ExcelHandle():
      #엑셀파일 파일읽기
      #path_dir = "C:"+"\\Users\\SKbroadband\\Downloads\\"
      file_list = os.listdir(data["dir_path"])
      file_list.sort()
      #파일이름에서 특정 string을 가지는 파일만 뽑아내기
      vocList = list()
      for item in file_list:
         if item.find('VoC상세내역') is not -1:
            vocList.append(item)
      #print(vocList)
      #try:
      for rfile in vocList:
         selected_file = data["dir_path"] + rfile

         wb_pyxl = load_workbook(selected_file)
         sheet = wb_pyxl.active #첫번째 쉬트를 활성화
         all_rows_count = sheet.max_row # 시트의 행수
         all_columns_count = sheet.max_column #시트의 컬럼수

         for rows in range(1,all_rows_count):   
            branch = sheet.cell(row=rows, column=2).value #본부
            catogory1 = sheet.cell(row=rows, column=9).value #IPTV장애
            catogory2 = sheet.cell(row=rows, column=10).value #IPTV_전체채널_끊김/모자이크
            #print(branch, catogory1, catogory2)
            messageTxt = ""
            if(branch ==data["branch"] and
               catogory1 =="IPTV장애" and
               (catogory2 == "IPTV_전체채널_끊김/모자이크" or
               catogory2 == "IPTV_특정채널_끊김/모자이크") 
               ):  
               print(branch, catogory1, catogory2)
               messageTxt = '[관리유통망]: ' + sheet.cell(row=rows, column=14).value + "\n"
               messageTxt = messageTxt + '[지역]: ' + sheet.cell(row=rows, column=30).value +" "+ sheet.cell(row=rows, column=31).value +" " +sheet.cell(row=rows, column=32).value + "\n"
               messageTxt = messageTxt + '[접수일]: ' + sheet.cell(row=rows, column=1).value + "\n"
               serviceNum = sheet.cell(row=rows, column=6).value 
               serviceNum = str(serviceNum)
               #messageTxt = messageTxt + str(sheet.cell(row=rows, column=6)).value + "\n"
               messageTxt = messageTxt + '[서비스번호]: ' + serviceNum + "\n"
               messageTxt = messageTxt + '[중분류]: ' + sheet.cell(row=rows, column=9).value + "\n"
               messageTxt = messageTxt + '[소분류]: ' + sheet.cell(row=rows, column=10).value + "\n"
               messageTxt = messageTxt + '[장비ITD]: ' + sheet.cell(row=rows, column=17).value + "\n"
               messageTxt = messageTxt + '[단말]: ' + sheet.cell(row=rows, column=20).value + "\n"
               messageTxt = messageTxt + '[공유기]: ' + sheet.cell(row=rows, column=22).value + "\n"
               messageTxt = messageTxt + '[STB]: ' + sheet.cell(row=rows, column=25).value + "\n"
               messageTxt = messageTxt + '[상담상세]\n' + sheet.cell(row=rows, column=35).value + "\n"
               telegramSend.telebot(messageTxt) #텔레그램으로 메세지를 보냅니다.

         os.remove(selected_file) 
         print(selected_file,"처리완료")
      #except:
      #print('PermissionError: [WinError 32] 다른 프로세스가 파일을 사용 중이기 때문에 프로세스가 액세스 할 수 없습니다:')

   ##########ADAMS 접속###############################
   driver = webdriver.Chrome(executable_path=data["cdrivefile"]) 
   driver.get("https://adams.skbroadband.com/") 

   time.sleep(5)
   driver.switch_to.frame("warpFrame") 

   elem = driver.find_element_by_id('loginid') 
   elem.send_keys(data["adams_userid"]) 

   elem = driver.find_element_by_id("pkey") 
   elem.send_keys(data["adams_password"]) 

   elem = driver.find_element_by_id("btnLogin").click() 
   #elem.submit() 
   time.sleep(8)
   window_current=driver.window_handles[0]

   ##########ADAMS VoC 상세화면 이동 및 검색#############
   driver.find_element_by_tag_name("body").send_keys(Keys.CONTROL + 't') #새Tab을 만드는 것인데, 동작하지 않음(확인이 필요함)
   #VoC상세내역 창으로 이동 함
   driver.get("https://adams.skbroadband.com/web/jsp/fm/PG-FM-6002.jsp")
   time.sleep(2)
   elem = driver.find_element_by_xpath('/html/body/div[2]/div[2]/div[1]/ul/li[2]').click()  
   time.sleep(1)
   searchEndTime =""
   ## 주기적으로 단위로 계속 자료를 검색한다.
   period = int(data["period"]) #5분주기로 처리함
   while True: 
         now = datetime.datetime.now()
         if searchEndTime=="":
            searchEndTime = now - datetime.timedelta(minutes=period)
         searchStartTime = now
         print(searchEndTime, "~", searchStartTime)	

         #검색기준 시작시간
         srch_sta_dt = searchEndTime.strftime('%Y-%m-%d')
         srch_sta_hr = searchEndTime.strftime('%H')
         srch_sta_min = searchEndTime.strftime('%M')
         elem = driver.find_element_by_id("srch_sta_dt") 
         elem.send_keys(srch_sta_dt) 
         sta_hr = Select(driver.find_element_by_id('srch_sta_hr'))
         sta_hr.select_by_value(srch_sta_hr)
         sta_min = Select(driver.find_element_by_id('srch_sta_min'))
         sta_min.select_by_value(srch_sta_min)
         #검색기준 끝시간
         srch_end_dt = searchStartTime.strftime('%Y-%m-%d')
         srch_end_hr = searchStartTime.strftime('%H')
         srch_end_min = searchStartTime.strftime('%M')
         elem = driver.find_element_by_id("srch_end_dt") 
         elem.send_keys(srch_end_dt) 
         end_hr = Select(driver.find_element_by_id('srch_end_hr'))
         end_hr.select_by_value(srch_end_hr)
         end_min = Select(driver.find_element_by_id('srch_end_min'))
         end_min.select_by_value(srch_end_min)
         #검색기준 본부선택 멀티선택기능으로 동작이 안됨(확인이 필요함)
         branch = Select(driver.find_element_by_id('srch_branch'))
         #branch.select_by_visible_text('서부Infra본부')
         branch.select_by_value('NB00700000')
         #Excel Download 클릭
         elem = driver.find_element_by_id("btnExcel").click() 
         time.sleep(8) #파일이 완전히 저장이 될때까지 기다린다.
         #기준시간을 이전시간으로 변경		 
         searchEndTime = now
		 #Chronic VoC가 발생하면 메세지를 보내기 위해 검사
         ExcelHandle()
         #검색주기를 설정 
         sleep_Time = period*60
         time.sleep(sleep_Time) # 검색주기동안 Sleep
   #브라우저를 종료
   driver.close()